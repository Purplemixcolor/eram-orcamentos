from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
from dataclasses import asdict, dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook

try:
    import pdfplumber
except Exception:  # pragma: no cover - optional at runtime
    pdfplumber = None


UNIT_TOKENS = {
    "UN",
    "UND",
    "UNID",
    "UNIDADE",
    "SV",
    "SERV",
    "SERVICO",
    "SERVIÇO",
    "M",
    "M2",
    "M²",
    "M3",
    "M³",
    "KG",
    "TON",
    "H",
    "HH",
    "H/H",
    "DIA",
    "VB",
    "CJ",
    "PÇ",
    "PC",
    "LT",
    "GL",
}

NON_BUDGET_NAME_HINTS = {
    "CALCULADORA",
    "IMPOSTOS",
    "STATUS",
    "REB",
    "FORNECEDORES",
    "ESPECIFICAÇÃO",
    "ESPECIFICACAO",
    "MODELO",
}

SKIP_WORDS = {
    "TOTAL",
    "SUBTOTAL",
    "IMPOSTO",
    "IMPOSTOS",
    "ISS",
    "INSS",
    "IRRF",
    "PIS",
    "COFINS",
    "CSLL",
    "DEDUCAO",
    "DEDUÇÃO",
    "RETENCOES",
    "RETENÇÕES",
    "VALOR LIQUIDO",
    "VALOR LÍQUIDO",
}


@dataclass
class CatalogItem:
    source_file: str
    source_type: str
    source_sheet: str
    source_page: str
    source_row: int | None
    extraction_pattern: str
    confidence: str
    year: int | None
    quote_number: str
    work_order_number: str
    shipowner: str
    vessel: str
    item_code: str
    title: str
    description: str
    unit: str
    quantity: float | None
    unit_value: float | None
    total_value: float | None
    currency: str
    category_hint: str
    review_notes: str


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\n", " ").replace("\r", " ")
    return re.sub(r"\s+", " ", text).strip()


def parse_number(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = normalize_text(value)
    if not text:
        return None
    text = text.replace("R$", "").replace(" ", "")
    text = re.sub(r"[^0-9,.\-]", "", text)
    if not text or text in {"-", ".", ","}:
        return None
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    elif "," in text:
        text = text.replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


def looks_like_code(value: str) -> bool:
    return bool(re.match(r"^\d+(\.\d+)*$", value)) or bool(re.match(r"^[A-Z]{1,4}\d{2,}", value))


def looks_like_unit(value: str) -> bool:
    return normalize_text(value).upper().replace(".", "") in UNIT_TOKENS


def should_skip_description(text: str) -> bool:
    upper = normalize_text(text).upper()
    return any(word in upper for word in SKIP_WORDS)


def should_skip_workbook(path: Path) -> bool:
    upper = str(path).upper()
    return any(hint in upper for hint in NON_BUDGET_NAME_HINTS)


def infer_year(path: Path, *texts: str) -> int | None:
    haystack = " ".join([str(path), *texts])
    years = [int(match) for match in re.findall(r"\b(20[12][0-9])\b", haystack)]
    valid = [year for year in years if 2010 <= year <= datetime.now().year + 1]
    return valid[-1] if valid else None


def infer_quote_number(path: Path, *texts: str) -> str:
    haystack = " ".join([path.stem, *texts])
    patterns = [
        r"\bMC\s*[-:]?\s*(\d{3,6})\b",
        r"\bOR[ÇC]AMENTO\s*(?:N[ºO.]*)?\s*[-:]?\s*([A-Z0-9./-]{2,})",
        r"\bORC[-\s]?(\d{2,6})\b",
    ]
    for pattern in patterns:
        match = re.search(pattern, haystack, flags=re.IGNORECASE)
        if match:
            return match.group(0).strip()
    return ""


def infer_entities(path: Path, *texts: str) -> tuple[str, str]:
    name = path.stem
    parts = re.split(r"\s+-\s+|\s+MC\s*\d+|_", name)
    vessel = parts[0].strip(" -") if parts else name
    shipowner = ""
    joined = " ".join(texts)
    ref = re.search(r"Ref:\s*([^\\n]+)", joined, flags=re.IGNORECASE)
    if ref:
        vessel = normalize_text(ref.group(1))[:120]
    company = re.search(r"\b(?:CNA|FOG[ÁA]S|NAVECUNHA|AMAGGI|TRANSDOURADA|WPL|VDA|CIDADE TRANSPORTES|RIO NEGRO)\b", str(path), flags=re.IGNORECASE)
    if company:
        shipowner = company.group(0).upper()
    return shipowner, vessel


def category_hint(text: str) -> str:
    upper = text.upper()
    rules = [
        ("Pintura e tratamento", ["PINTURA", "JATEAMENTO", "INTERTUF", "TINTA"]),
        ("Casco", ["CASCO", "CHAPEAMENTO", "CHAPA", "BOJO", "COSTADO"]),
        ("Soldagem", ["SOLDA", "SOLDAGEM", "ELETRODO", "ARAME TUBULAR"]),
        ("Sistema eletrico", ["ELETRIC", "CABO", "LUMINARIA", "REFLETOR", "QUADRO"]),
        ("Sistema de propulsao", ["HELICE", "HÉLICE", "EIXO PROPULSOR", "PROPULSAO"]),
        ("Sistema hidraulico", ["HIDRAULIC", "CILINDRO", "BOMBA"]),
        ("Docagem", ["DOCAGEM", "DESDOCAGEM", "DIQUE", "CARREIRA"]),
        ("Tubulacao", ["TUBO", "TUBULACAO", "TUBULAÇÃO", "VALVULA", "VÁLVULA"]),
        ("Estruturas metalicas", ["ESTRUTURA", "CANTONEIRA", "VIGA", "SUPORTE"]),
    ]
    for category, needles in rules:
        if any(needle in upper for needle in needles):
            return category
    return "Servicos gerais"


def row_values(row: Iterable[Any]) -> list[Any]:
    values = list(row)
    while values and values[-1] in (None, ""):
        values.pop()
    return values


def extract_itemized_rows(path: Path, sheet_name: str, rows: list[list[Any]]) -> list[CatalogItem]:
    items: list[CatalogItem] = []
    sheet_context = " ".join(normalize_text(cell) for row in rows[:12] for cell in row[:8])
    shipowner, vessel = infer_entities(path, sheet_context)
    year = infer_year(path, sheet_context)
    quote_number = infer_quote_number(path, sheet_context)

    for index, row in enumerate(rows, start=1):
        values = row_values(row)
        if len(values) < 4:
            continue

        strings = [normalize_text(value) for value in values]
        code_idx = next((i for i, value in enumerate(strings[:3]) if looks_like_code(value)), None)
        if code_idx is None:
            continue

        desc_idx = None
        for i in range(code_idx + 1, min(len(values), code_idx + 5)):
            text = strings[i]
            if len(text) >= 8 and not looks_like_unit(text) and parse_number(text) is None:
                desc_idx = i
                break
        if desc_idx is None:
            continue

        description = strings[desc_idx]
        if should_skip_description(description):
            continue

        unit = ""
        unit_idx = None
        for i in range(desc_idx + 1, min(len(values), desc_idx + 4)):
            if looks_like_unit(strings[i]):
                unit = strings[i].upper()
                unit_idx = i
                break

        numeric_cells: list[tuple[int, float]] = []
        for i in range((unit_idx or desc_idx) + 1, len(values)):
            number = parse_number(values[i])
            if number is not None:
                numeric_cells.append((i, number))

        if len(numeric_cells) < 2:
            continue

        quantity = numeric_cells[0][1]
        unit_value = numeric_cells[1][1] if unit_idx is not None and len(numeric_cells) >= 2 else numeric_cells[-2][1]
        total_value = numeric_cells[2][1] if unit_idx is not None and len(numeric_cells) >= 3 else numeric_cells[-1][1]
        if total_value <= 0:
            continue

        if unit_value and quantity and abs((quantity * unit_value) - total_value) / max(total_value, 1) > 0.35:
            confidence = "Media"
            review_notes = "Valores extraidos por posicao; conferir quantidade, unitario e total."
        else:
            confidence = "Alta"
            review_notes = ""

        items.append(
            CatalogItem(
                source_file=str(path),
                source_type=path.suffix.lower().lstrip("."),
                source_sheet=sheet_name,
                source_page="",
                source_row=index,
                extraction_pattern="spreadsheet_itemized",
                confidence=confidence,
                year=year,
                quote_number=quote_number,
                work_order_number="",
                shipowner=shipowner,
                vessel=vessel,
                item_code=strings[code_idx],
                title=description[:140],
                description=description,
                unit=unit,
                quantity=quantity,
                unit_value=unit_value,
                total_value=total_value,
                currency="BRL",
                category_hint=category_hint(description),
                review_notes=review_notes,
            )
        )
    return items


def extract_transposed_price_table(path: Path, sheet_name: str, rows: list[list[Any]]) -> list[CatalogItem]:
    items: list[CatalogItem] = []
    for header_index, row in enumerate(rows, start=1):
        headers = [normalize_text(cell) for cell in row]
        upper = [value.upper() for value in headers]
        if not (any("ARM" in value for value in upper[:5]) and any(value == "MC" for value in upper[:8])):
            continue
        service_start = 3
        for data_index, data_row in enumerate(rows[header_index:], start=header_index + 1):
            if len(data_row) < service_start + 1:
                continue
            shipowner = normalize_text(data_row[0])
            quote_number = normalize_text(data_row[1])
            vessel = normalize_text(data_row[2])
            if not shipowner and not quote_number and not vessel:
                continue
            for col in range(service_start, min(len(headers), len(data_row))):
                title = headers[col]
                value = parse_number(data_row[col])
                if not title or value is None or value <= 0 or should_skip_description(title):
                    continue
                items.append(
                    CatalogItem(
                        source_file=str(path),
                        source_type=path.suffix.lower().lstrip("."),
                        source_sheet=sheet_name,
                        source_page="",
                        source_row=data_index,
                        extraction_pattern="spreadsheet_transposed_price",
                        confidence="Media",
                        year=infer_year(path, sheet_name, title),
                        quote_number=quote_number,
                        work_order_number="",
                        shipowner=shipowner,
                        vessel=vessel,
                        item_code="",
                        title=title[:140],
                        description=title,
                        unit="",
                        quantity=1,
                        unit_value=value,
                        total_value=value,
                        currency="BRL",
                        category_hint=category_hint(title),
                        review_notes="Tabela de precos transposta; unidade e quantidade assumidas como 1 para referencia.",
                    )
                )
        break
    return items


def extract_workbook(path: Path, max_rows: int = 5000, max_cols: int = 80) -> tuple[list[CatalogItem], str]:
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        return [], f"Erro ao abrir planilha: {type(exc).__name__}: {exc}"

    extracted: list[CatalogItem] = []
    try:
        for sheet_name in workbook.sheetnames:
            ws = workbook[sheet_name]
            if not hasattr(ws, "iter_rows"):
                continue
            row_limit = min(ws.max_row or 0, max_rows)
            col_limit = min(ws.max_column or 0, max_cols)
            rows = [row_values(row) for row in ws.iter_rows(max_row=row_limit, max_col=col_limit, values_only=True)]
            extracted.extend(extract_itemized_rows(path, sheet_name, rows))
            extracted.extend(extract_transposed_price_table(path, sheet_name, rows))
    finally:
        workbook.close()
    return extracted, ""


def extract_pdf(path: Path, max_pages: int | None = None) -> tuple[list[CatalogItem], str]:
    if pdfplumber is None:
        return [], "pdfplumber indisponivel"
    extracted: list[CatalogItem] = []
    try:
        with pdfplumber.open(path) as pdf:
            pages = pdf.pages if max_pages is None else pdf.pages[:max_pages]
            first_text = pages[0].extract_text() if pages else ""
            shipowner, vessel = infer_entities(path, first_text or "")
            quote_number = infer_quote_number(path, first_text or "")
            year = infer_year(path, first_text or "")
            for page_number, page in enumerate(pages, start=1):
                text = page.extract_text() or ""
                lines = [normalize_text(line) for line in text.splitlines() if normalize_text(line)]
                for line_index, line in enumerate(lines):
                    item_match = re.match(r"^(\d+(?:\.\d+)+)\s+(.+)$", line)
                    if item_match and not should_skip_description(line):
                        numbers = re.findall(r"(?:R\$\s*)?-?\d[\d. ]*,\d{2}", line)
                        unit_match = re.search(r"\b(" + "|".join(re.escape(unit) for unit in sorted(UNIT_TOKENS, key=len, reverse=True)) + r")\b", line, flags=re.IGNORECASE)
                        if len(numbers) >= 2 and unit_match:
                            unit_value = parse_number(numbers[-2])
                            total_value = parse_number(numbers[-1])
                            description = re.sub(r"(?:R\$\s*)?-?\d[\d. ]*,\d{2}", "", item_match.group(2))
                            description = re.sub(r"\b" + re.escape(unit_match.group(1)) + r"\b", "", description, flags=re.IGNORECASE)
                            description = normalize_text(description)
                            if len(description) < 8 and line_index > 0:
                                description = normalize_text(lines[line_index - 1] + " " + description)
                            if total_value and total_value > 0:
                                extracted.append(
                                    CatalogItem(
                                        source_file=str(path),
                                        source_type="pdf",
                                        source_sheet="",
                                        source_page=str(page_number),
                                        source_row=None,
                                        extraction_pattern="pdf_budget_line",
                                        confidence="Baixa",
                                        year=year,
                                        quote_number=quote_number,
                                        work_order_number="",
                                        shipowner=shipowner,
                                        vessel=vessel,
                                        item_code=item_match.group(1),
                                        title=description[:140],
                                        description=description,
                                        unit=unit_match.group(1).upper(),
                                        quantity=None,
                                        unit_value=unit_value,
                                        total_value=total_value,
                                        currency="BRL",
                                        category_hint=category_hint(description),
                                        review_notes="Extraido de texto PDF; revisar contra o documento original.",
                                    )
                                )

                    material_match = re.match(r"^(\d{3,6}):\s+(.+?)\s+(\d{4,8})\s+(\d{2}/\d{2}/\d{2})\s+([\d.,]+)\s+([\d.,]+)\s+([\d.,]+)$", line)
                    if material_match:
                        description = normalize_text(material_match.group(2))
                        quantity = parse_number(material_match.group(5))
                        unit_value = parse_number(material_match.group(6))
                        total_value = parse_number(material_match.group(7))
                        if description and total_value and total_value > 0:
                            extracted.append(
                                CatalogItem(
                                    source_file=str(path),
                                    source_type="pdf",
                                    source_sheet="",
                                    source_page=str(page_number),
                                    source_row=None,
                                    extraction_pattern="pdf_cost_center_material",
                                    confidence="Baixa",
                                    year=year,
                                    quote_number=quote_number,
                                    work_order_number=material_match.group(3),
                                    shipowner=shipowner,
                                    vessel=vessel,
                                    item_code=material_match.group(1),
                                    title=description[:140],
                                    description=description,
                                    unit="",
                                    quantity=quantity,
                                    unit_value=unit_value,
                                    total_value=total_value,
                                    currency="BRL",
                                    category_hint=category_hint(description),
                                    review_notes="Item de centro de custo extraido de PDF; confirmar se deve compor catalogo oficial de servicos ou materiais.",
                                )
                            )
    except Exception as exc:
        return extracted, f"Erro ao ler PDF: {type(exc).__name__}: {exc}"
    return extracted, ""


def file_sha1(path: Path) -> str:
    digest = hashlib.sha1()
    with path.open("rb") as file:
        for chunk in iter(lambda: file.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def write_csv(path: Path, rows: list[dict[str, Any]], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def main() -> None:
    parser = argparse.ArgumentParser(description="Extrai catalogo historico de precos dos arquivos de orçamento ERAM.")
    parser.add_argument("source", type=Path, help="Pasta raiz dos orcamentos")
    parser.add_argument("--output", type=Path, default=Path("private-data/catalog"))
    parser.add_argument("--include-pdf", action="store_true", help="Tambem tenta extrair linhas de PDFs")
    parser.add_argument("--pdf-max-pages", type=int, default=None, help="Limite de paginas por PDF")
    parser.add_argument("--extensions", default="", help="Lista de extensoes para processar/inventariar, ex: .xlsx,.pdf")
    parser.add_argument("--hash-files", action="store_true", help="Calcula SHA1 dos arquivos inventariados")
    args = parser.parse_args()

    source = args.source
    if not source.exists():
        raise SystemExit(f"Pasta nao encontrada: {source}")

    extension_filter = {item.strip().lower() for item in args.extensions.split(",") if item.strip()}
    files = [path for path in source.rglob("*") if path.is_file() and (not extension_filter or path.suffix.lower() in extension_filter)]
    inventory: list[dict[str, Any]] = []
    items: list[CatalogItem] = []
    errors: list[dict[str, Any]] = []

    for counter, path in enumerate(files, start=1):
        suffix = path.suffix.lower()
        rel = str(path.relative_to(source))
        stat = path.stat()
        inventory.append(
            {
                "relative_path": rel,
                "full_path": str(path),
                "extension": suffix,
                "size_bytes": stat.st_size,
                "modified_at": datetime.fromtimestamp(stat.st_mtime).isoformat(timespec="seconds"),
                "sha1": file_sha1(path) if args.hash_files and stat.st_size <= 20 * 1024 * 1024 else "",
            }
        )

        extracted: list[CatalogItem] = []
        error = ""
        if suffix == ".xlsx":
            if should_skip_workbook(path):
                error = "Planilha administrativa/modelo inventariada, mas excluida do catalogo."
            else:
                extracted, error = extract_workbook(path)
        elif suffix == ".pdf" and args.include_pdf:
            extracted, error = extract_pdf(path, args.pdf_max_pages)
        elif suffix in {".xls", ".ods"}:
            error = f"Formato {suffix} inventariado, mas nao extraido nesta etapa."

        if extracted:
            items.extend(extracted)
        if error:
            errors.append({"source_file": str(path), "extension": suffix, "error": error})

        if counter % 25 == 0:
            print(f"Processados {counter}/{len(files)} arquivos; itens extraidos: {len(items)}", flush=True)

    output = args.output
    item_rows = [asdict(item) for item in items]
    item_fields = list(asdict(CatalogItem("", "", "", "", None, "", "", None, "", "", "", "", "", "", "", "", None, None, None, "", "", "")).keys())
    write_csv(output / "catalog_services.csv", item_rows, item_fields)
    (output / "catalog_services.json").write_text(json.dumps(item_rows, ensure_ascii=False, indent=2), encoding="utf-8")
    write_csv(output / "source_inventory.csv", inventory, ["relative_path", "full_path", "extension", "size_bytes", "modified_at", "sha1"])
    write_csv(output / "extraction_errors.csv", errors, ["source_file", "extension", "error"])

    summary = {
        "source": str(source),
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "files_total": len(files),
        "items_total": len(items),
        "by_extension": {},
        "by_pattern": {},
        "by_confidence": {},
        "errors_total": len(errors),
    }
    for row in inventory:
        summary["by_extension"][row["extension"]] = summary["by_extension"].get(row["extension"], 0) + 1
    for item in items:
        summary["by_pattern"][item.extraction_pattern] = summary["by_pattern"].get(item.extraction_pattern, 0) + 1
        summary["by_confidence"][item.confidence] = summary["by_confidence"].get(item.confidence, 0) + 1
    (output / "extraction_summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
